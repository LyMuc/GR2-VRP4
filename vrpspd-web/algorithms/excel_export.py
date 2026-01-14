"""
Excel Export Module for VRPSPD
Xuất kết quả ra file Excel với formatting đẹp
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def create_excel_report(results, problem_info, filename=None):
    """
    Tạo file Excel báo cáo kết quả VRPSPD.
    
    Args:
        results (dict): Kết quả từ solve API
        problem_info (dict): Thông tin bài toán
        filename (str): Tên file (optional)
        
    Returns:
        str: Đường dẫn file Excel đã tạo
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"VRPSPD_Results_{timestamp}.xlsx"
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Summary
    create_summary_sheet(wb, results, problem_info)
    
    # Sheet 2: Route Details (best result)
    best_result = results.get('vnd') or results.get('savings')
    if best_result:
        create_route_details_sheet(wb, best_result, problem_info)
    
    # Sheet 3: Comparison (if both algorithms used)
    if 'savings' in results and 'vnd' in results:
        create_comparison_sheet(wb, results)
    
    # Save file
    filepath = os.path.join('static', 'uploads', filename)
    wb.save(filepath)
    
    return filepath


def create_summary_sheet(wb, results, problem_info):
    """Tạo sheet Summary với thông tin tổng hợp."""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = 'VRPSPD SOLUTION REPORT'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    # Timestamp
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:D2')
    
    # Problem Information
    row = 4
    ws[f'A{row}'] = 'PROBLEM INFORMATION'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    info_data = [
        ['Number of Customers:', problem_info.get('num_customers', '-')],
        ['Number of Vehicles:', problem_info.get('num_vehicles', '-')],
        ['Vehicle Capacity:', problem_info.get('capacity', '-')]
    ]
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # Results Summary
    row += 1
    ws[f'A{row}'] = 'RESULTS SUMMARY'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    # Header
    headers = ['Algorithm', 'Total Cost', 'Num Routes', 'Time (s)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    # Savings result
    if 'savings' in results:
        savings = results['savings']
        ws[f'A{row}'] = 'Savings-based'
        ws[f'B{row}'] = savings.get('total_cost', 0)
        ws[f'C{row}'] = savings.get('num_routes', 0)
        ws[f'D{row}'] = savings.get('computation_time', 0)
        row += 1
    
    # VND result
    if 'vnd' in results:
        vnd = results['vnd']
        ws[f'A{row}'] = 'Savings + VND'
        ws[f'B{row}'] = vnd.get('total_cost', 0)
        ws[f'C{row}'] = vnd.get('num_routes', 0)
        ws[f'D{row}'] = vnd.get('computation_time', 0)
        
        # Highlight VND row (best result)
        for col in range(1, 5):
            ws.cell(row, col).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            ws.cell(row, col).font = Font(bold=True)
        row += 1
    
    # Improvement
    if 'savings' in results and 'vnd' in results:
        row += 1
        ws[f'A{row}'] = 'Improvement:'
        ws[f'A{row}'].font = Font(bold=True)
        improvement = results['vnd'].get('improvement', 0)
        ws[f'B{row}'] = f"{improvement}%"
        ws[f'B{row}'].font = Font(bold=True, color='006100')
        ws[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # Apply borders
    apply_borders(ws, 1, 1, row, 4)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_route_details_sheet(wb, result, problem_info):
    """Tạo sheet chi tiết các tuyến."""
    ws = wb.create_sheet("Route Details")
    
    # Title
    ws['A1'] = 'ROUTE DETAILS'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 25
    
    # Headers
    row = 3
    headers = ['Route #', 'Sequence', 'Distance', 'Total Delivery', 'Total Pickup', 'Feasible']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row += 1
    for route_detail in result.get('route_details', []):
        ws[f'A{row}'] = route_detail['route_number']
        ws[f'B{row}'] = ' → '.join(map(str, route_detail['route_with_depot']))
        ws[f'C{row}'] = route_detail['distance']
        ws[f'D{row}'] = route_detail['total_delivery']
        ws[f'E{row}'] = route_detail['total_pickup']
        ws[f'F{row}'] = 'Yes'
        
        # Alternate row colors
        if row % 2 == 0:
            for col in range(1, 7):
                ws.cell(row, col).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        row += 1
    
    # Total row
    ws[f'A{row}'] = 'TOTAL'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = result.get('total_cost', 0)
    ws[f'C{row}'].font = Font(bold=True)
    
    for col in range(1, 7):
        ws.cell(row, col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # Apply borders
    apply_borders(ws, 3, 1, row, 6)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10


def create_comparison_sheet(wb, results):
    """Tạo sheet so sánh giữa Savings và VND."""
    ws = wb.create_sheet("Comparison")
    
    # Title
    ws['A1'] = 'ALGORITHM COMPARISON'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 25
    
    # Headers
    row = 3
    headers = ['Metric', 'Savings-based', 'Savings + VND', 'Difference', 'Improvement (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    savings = results['savings']
    vnd = results['vnd']
    
    row += 1
    # Total Cost
    ws[f'A{row}'] = 'Total Cost'
    ws[f'B{row}'] = savings['total_cost']
    ws[f'C{row}'] = vnd['total_cost']
    ws[f'D{row}'] = savings['total_cost'] - vnd['total_cost']
    ws[f'E{row}'] = vnd.get('improvement', 0)
    ws[f'E{row}'].number_format = '0.00"%"'
    
    # Highlight improvement
    for col in range(4, 6):
        ws.cell(row, col).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        ws.cell(row, col).font = Font(bold=True, color='006100')
    
    row += 1
    # Number of Routes
    ws[f'A{row}'] = 'Number of Routes'
    ws[f'B{row}'] = savings['num_routes']
    ws[f'C{row}'] = vnd['num_routes']
    ws[f'D{row}'] = savings['num_routes'] - vnd['num_routes']
    ws[f'E{row}'] = ''
    
    row += 1
    # Computation Time
    ws[f'A{row}'] = 'Computation Time (s)'
    ws[f'B{row}'] = savings['computation_time']
    ws[f'C{row}'] = vnd['computation_time']
    ws[f'D{row}'] = vnd['computation_time'] - savings['computation_time']
    ws[f'E{row}'] = ''
    
    # Apply borders
    apply_borders(ws, 3, 1, row, 5)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18


def apply_borders(ws, start_row, start_col, end_row, end_col):
    """Áp dụng borders cho một vùng cells."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = thin_border
            if ws.cell(row, col).value is not None:
                ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
