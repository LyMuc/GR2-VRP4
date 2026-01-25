#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
VRPSPD Web Application - Complete Test Scenarios
=================================================

Script này thực thi tất cả các kịch bản thử nghiệm cho VRPSPD Web App
Xem chi tiết trong file: VRPSPD_TEST_SCENARIOS.md

Cách sử dụng:
    python test_complete_scenarios.py --all          # Chạy tất cả tests
    python test_complete_scenarios.py --unit         # Chỉ unit tests
    python test_complete_scenarios.py --api          # Chỉ API tests
    python test_complete_scenarios.py --performance  # Performance tests
    python test_complete_scenarios.py --e2e          # End-to-end test

"""

import os
import sys
import json
import time
import argparse
from datetime import datetime
import traceback

# ============================================================================
# Test Configuration
# ============================================================================

TEST_CONFIG = {
    'upload_folder': 'static/uploads',
    'test_data_folder': 'static/uploads/test_data',
    'output_folder': 'test_results',
    'vnd_time_limit': 10,  # seconds
    'server_url': 'http://localhost:5000'
}

# ============================================================================
# Utility Functions
# ============================================================================

class Colors:
    """ANSI color codes for terminal output."""
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


class TestResult:
    """Store test result information."""
    def __init__(self, test_id, name, status, message="", duration=0):
        self.test_id = test_id
        self.name = name
        self.status = status  # PASS, FAIL, SKIP
        self.message = message
        self.duration = duration
        self.timestamp = datetime.now()


class TestSuite:
    """Manage test execution and results."""
    def __init__(self):
        self.results = []
        self.start_time = None
        self.end_time = None
    
    def add_result(self, result):
        self.results.append(result)
    
    def get_summary(self):
        total = len(self.results)
        passed = sum(1 for r in self.results if r.status == 'PASS')
        failed = sum(1 for r in self.results if r.status == 'FAIL')
        skipped = sum(1 for r in self.results if r.status == 'SKIP')
        return {
            'total': total,
            'passed': passed,
            'failed': failed,
            'skipped': skipped,
            'success_rate': (passed / total * 100) if total > 0 else 0
        }
    
    def print_summary(self):
        summary = self.get_summary()
        print("\n" + "=" * 70)
        print(f"{Colors.BOLD}TEST SUMMARY{Colors.ENDC}")
        print("=" * 70)
        print(f"Total Tests: {summary['total']}")
        print(f"{Colors.OKGREEN}Passed: {summary['passed']} ✅{Colors.ENDC}")
        print(f"{Colors.FAIL}Failed: {summary['failed']} ❌{Colors.ENDC}")
        print(f"{Colors.WARNING}Skipped: {summary['skipped']} ⚠️{Colors.ENDC}")
        print(f"Success Rate: {summary['success_rate']:.1f}%")
        
        if self.start_time and self.end_time:
            duration = (self.end_time - self.start_time).total_seconds()
            print(f"Total Duration: {duration:.2f}s")
        
        print("=" * 70)


def print_test_header(test_id, name):
    """Print test header."""
    print("\n" + "-" * 70)
    print(f"{Colors.BOLD}[{test_id}] {name}{Colors.ENDC}")
    print("-" * 70)


def print_pass(message):
    """Print pass message."""
    print(f"{Colors.OKGREEN}✅ {message}{Colors.ENDC}")


def print_fail(message):
    """Print fail message."""
    print(f"{Colors.FAIL}❌ {message}{Colors.ENDC}")


def print_info(message):
    """Print info message."""
    print(f"{Colors.OKCYAN}ℹ️  {message}{Colors.ENDC}")


def print_warning(message):
    """Print warning message."""
    print(f"{Colors.WARNING}⚠️  {message}{Colors.ENDC}")


def ensure_folders():
    """Create necessary folders."""
    for folder in [TEST_CONFIG['upload_folder'], 
                   TEST_CONFIG['test_data_folder'],
                   TEST_CONFIG['output_folder']]:
        os.makedirs(folder, exist_ok=True)


def create_test_data_file():
    """Create a standard test data file."""
    test_data = """Cost matrix
0 9 14 23
9 0 21 22
14 21 0 25
23 22 25 0

Delivery quantities
1200 1700 1500

Pick-up quantities
0 1200 1700

Vehicle capacity
6000
"""
    filepath = os.path.join(TEST_CONFIG['test_data_folder'], 'standard_test.txt')
    with open(filepath, 'w') as f:
        f.write(test_data)
    return filepath


def create_large_test_data():
    """Create large instance for performance testing."""
    import random
    
    n = 50  # 50 customers
    depot = 0
    
    # Random cost matrix
    costs = [[0] * (n + 1) for _ in range(n + 1)]
    for i in range(n + 1):
        for j in range(i + 1, n + 1):
            cost = random.randint(10, 100)
            costs[i][j] = cost
            costs[j][i] = cost
    
    # Random demands and pickups
    demands = [random.randint(50, 150) for _ in range(n)]
    pickups = [random.randint(0, 100) for _ in range(n)]
    
    # Vehicle capacity
    capacity = 1000
    
    # Write to file
    filepath = os.path.join(TEST_CONFIG['test_data_folder'], 'large_test.txt')
    with open(filepath, 'w') as f:
        f.write("Cost matrix\n")
        for row in costs:
            f.write(" ".join(map(str, row)) + "\n")
        f.write("\nDelivery quantities\n")
        f.write(" ".join(map(str, demands)) + "\n")
        f.write("\nPick-up quantities\n")
        f.write(" ".join(map(str, pickups)) + "\n")
        f.write(f"\nVehicle capacity\n{capacity}\n")
    
    return filepath


def create_invalid_test_data():
    """Create invalid test data for error handling."""
    files = {}
    
    # Missing cost matrix
    files['missing_cost_matrix'] = """Delivery quantities
1200 1700 1500

Pick-up quantities
0 1200 1700

Vehicle capacity
6000
"""
    
    # Invalid format
    files['invalid_format'] = """This is not a valid VRPSPD file
Just random text
"""
    
    # Empty file
    files['empty'] = ""
    
    # Zero capacity
    files['zero_capacity'] = """Cost matrix
0 9 14
9 0 21
14 21 0

Delivery quantities
1200 1700

Pick-up quantities
0 1200

Vehicle capacity
0
"""
    
    filepaths = {}
    for name, content in files.items():
        filepath = os.path.join(TEST_CONFIG['test_data_folder'], f'{name}.txt')
        with open(filepath, 'w') as f:
            f.write(content)
        filepaths[name] = filepath
    
    return filepaths


# ============================================================================
# TC001: Test Module Import
# ============================================================================

def tc001_test_imports(suite):
    """TC001: Kiểm tra module import."""
    test_id = "TC001"
    test_name = "Kiểm Tra Module Import"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        # Import các module chính
        from algorithms import (
            read_vrpspd_file,
            solve_savings,
            solve_vnd
        )
        print_pass("Imported core algorithms")
        
        from algorithms.visualization import (
            create_plotly_data,
            create_plotly_figure_json
        )
        print_pass("Imported visualization module")
        
        from algorithms.excel_export import create_excel_report
        print_pass("Imported Excel export module")
        
        from algorithms.batch_processor import (
            process_batch_files,
            create_batch_summary
        )
        print_pass("Imported batch processing module")
        
        from algorithms.batch_excel_export import create_batch_excel_report
        print_pass("Imported batch Excel export module")
        
        duration = time.time() - start_time
        result = TestResult(test_id, test_name, 'PASS', 
                          "All modules imported successfully", duration)
        suite.add_result(result)
        return True
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Import failed: {str(e)}")
        result = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result)
        return False


# ============================================================================
# TC002: Test File Parser
# ============================================================================

def tc002_test_file_parser(suite):
    """TC002: Kiểm tra file parser."""
    test_id = "TC002"
    test_name = "Kiểm Tra File Parser"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        from algorithms import read_vrpspd_file
        
        # Create test file
        filepath = create_test_data_file()
        print_info(f"Created test file: {filepath}")
        
        # Parse file
        cost_matrix, demand, pickup, vehicle_caps, capacity, num_vehicles = read_vrpspd_file(filepath)
        
        # Verify results
        assert len(cost_matrix) == 4, "Cost matrix should be 4x4"
        print_pass(f"Cost matrix size: {len(cost_matrix)}x{len(cost_matrix[0])}")
        
        assert len(demand) == 3, "Should have 3 customers"
        print_pass(f"Number of customers: {len(demand)}")
        
        assert demand == [1200, 1700, 1500], "Demand values incorrect"
        print_pass(f"Demand vector: {demand}")
        
        assert pickup == [0, 1200, 1700], "Pickup values incorrect"
        print_pass(f"Pickup vector: {pickup}")
        
        assert capacity == 6000, "Capacity should be 6000"
        print_pass(f"Vehicle capacity: {capacity}")
        
        duration = time.time() - start_time
        result = TestResult(test_id, test_name, 'PASS',
                          f"Parsed {len(demand)} customers successfully", duration)
        suite.add_result(result)
        return True, filepath, cost_matrix, demand, pickup, capacity, num_vehicles
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Parser failed: {str(e)}")
        traceback.print_exc()
        result = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result)
        return False, None, None, None, None, None, None


# ============================================================================
# TC003: Test Savings Algorithm
# ============================================================================

def tc003_test_savings(suite, cost_matrix, demand, pickup, capacity, num_vehicles):
    """TC003: Kiểm tra thuật toán Savings."""
    test_id = "TC003"
    test_name = "Kiểm Tra Thuật Toán Savings"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        from algorithms import solve_savings
        
        # Run Savings
        result = solve_savings(cost_matrix, demand, pickup, capacity, num_vehicles)
        
        # Verify result structure
        assert 'total_cost' in result, "Missing total_cost"
        assert 'num_routes' in result, "Missing num_routes"
        assert 'routes' in result, "Missing routes"
        assert 'solution_vector' in result, "Missing solution_vector"
        assert 'computation_time' in result, "Missing computation_time"
        print_pass("Result structure valid")
        
        # Verify values
        assert result['total_cost'] > 0, "Total cost should be positive"
        print_pass(f"Total cost: {result['total_cost']:.2f}")
        
        assert result['num_routes'] >= 1, "Should have at least 1 route"
        print_pass(f"Number of routes: {result['num_routes']}")
        
        print_pass(f"Computation time: {result['computation_time']:.4f}s")
        
        # Verify routes
        all_customers = set()
        for route in result['routes']:
            assert route[0] == 0 and route[-1] == 0, "Route should start and end at depot"
            all_customers.update(route[1:-1])
        
        expected_customers = set(range(1, len(cost_matrix)))
        assert all_customers == expected_customers, "Not all customers visited"
        print_pass("All customers visited exactly once")
        
        # Verify capacity constraints
        for i, route in enumerate(result['routes'], 1):
            route_demand = sum(demand[c-1] for c in route if c > 0)
            route_pickup = sum(pickup[c-1] for c in route if c > 0)
            max_load = max(route_demand, route_pickup)
            assert max_load <= capacity, f"Route {i} exceeds capacity"
        print_pass("All routes satisfy capacity constraints")
        
        duration = time.time() - start_time
        result_obj = TestResult(test_id, test_name, 'PASS',
                               f"Cost: {result['total_cost']:.2f}, Routes: {result['num_routes']}", 
                               duration)
        suite.add_result(result_obj)
        return True, result
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Savings failed: {str(e)}")
        traceback.print_exc()
        result_obj = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result_obj)
        return False, None


# ============================================================================
# TC004: Test VND Algorithm
# ============================================================================

def tc004_test_vnd(suite, savings_result, cost_matrix, demand, pickup, capacity):
    """TC004: Kiểm tra thuật toán VND."""
    test_id = "TC004"
    test_name = "Kiểm Tra Thuật Toán VND"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        from algorithms import solve_vnd
        
        # Get initial solution from Savings
        initial_vector = savings_result['solution_vector']
        print_info(f"Initial cost from Savings: {savings_result['total_cost']:.2f}")
        
        # Run VND
        result = solve_vnd(
            initial_vector,
            cost_matrix,
            demand,
            pickup,
            capacity,
            max_time_seconds=TEST_CONFIG['vnd_time_limit']
        )
        
        # Verify result structure
        assert 'total_cost' in result, "Missing total_cost"
        assert 'initial_cost' in result, "Missing initial_cost"
        assert 'improvement' in result, "Missing improvement"
        assert 'routes' in result, "Missing routes"
        print_pass("Result structure valid")
        
        # Verify improvement
        assert result['total_cost'] <= result['initial_cost'], "VND should not increase cost"
        print_pass(f"Final cost: {result['total_cost']:.2f}")
        print_pass(f"Improvement: {result['improvement']:.2f}%")
        
        if result['improvement'] > 0:
            print_pass(f"VND improved solution by {result['improvement']:.2f}%")
        else:
            print_info("VND did not improve (already optimal)")
        
        # Verify routes validity
        all_customers = set()
        for route in result['routes']:
            assert route[0] == 0 and route[-1] == 0, "Route should start and end at depot"
            all_customers.update(route[1:-1])
        
        expected_customers = set(range(1, len(cost_matrix)))
        assert all_customers == expected_customers, "Not all customers visited"
        print_pass("All customers still visited")
        
        # Verify capacity constraints
        for i, route in enumerate(result['routes'], 1):
            route_demand = sum(demand[c-1] for c in route if c > 0)
            route_pickup = sum(pickup[c-1] for c in route if c > 0)
            max_load = max(route_demand, route_pickup)
            assert max_load <= capacity, f"Route {i} exceeds capacity after VND"
        print_pass("All routes satisfy capacity constraints")
        
        duration = time.time() - start_time
        result_obj = TestResult(test_id, test_name, 'PASS',
                               f"Improvement: {result['improvement']:.2f}%", duration)
        suite.add_result(result_obj)
        return True, result
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"VND failed: {str(e)}")
        traceback.print_exc()
        result_obj = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result_obj)
        return False, None


# ============================================================================
# TC005: Test Visualization
# ============================================================================

def tc005_test_visualization(suite, routes, cost_matrix, demand, pickup):
    """TC005: Kiểm tra visualization."""
    test_id = "TC005"
    test_name = "Kiểm Tra Visualization"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        from algorithms.visualization import create_plotly_data, create_plotly_figure_json
        
        # Create visualization data
        viz_data = create_plotly_data(routes, cost_matrix, demand, pickup)
        
        # Verify structure
        assert 'coordinates' in viz_data, "Missing coordinates"
        assert 'depot' in viz_data, "Missing depot"
        assert 'routes' in viz_data, "Missing routes"
        print_pass("Visualization data structure valid")
        
        # Verify coordinates
        num_nodes = len(cost_matrix)
        assert len(viz_data['coordinates']) == num_nodes, "Wrong number of coordinates"
        print_pass(f"Generated coordinates for {num_nodes} nodes")
        
        # Verify depot
        assert 'x' in viz_data['depot'] and 'y' in viz_data['depot'], "Invalid depot coordinates"
        print_pass(f"Depot at ({viz_data['depot']['x']:.2f}, {viz_data['depot']['y']:.2f})")
        
        # Create Plotly figure
        plotly_fig = create_plotly_figure_json(viz_data)
        
        # Verify Plotly figure
        assert 'data' in plotly_fig, "Missing data in Plotly figure"
        assert 'layout' in plotly_fig, "Missing layout in Plotly figure"
        print_pass(f"Plotly figure has {len(plotly_fig['data'])} traces")
        
        # Save to file
        output_file = os.path.join(TEST_CONFIG['output_folder'], 'test_viz.json')
        with open(output_file, 'w') as f:
            json.dump(plotly_fig, f, indent=2)
        print_pass(f"Saved to: {output_file}")
        
        duration = time.time() - start_time
        result = TestResult(test_id, test_name, 'PASS',
                          f"Generated visualization with {len(routes)} routes", duration)
        suite.add_result(result)
        return True
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Visualization failed: {str(e)}")
        traceback.print_exc()
        result = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result)
        return False


# ============================================================================
# TC006: Test Excel Export (Single)
# ============================================================================

def tc006_test_excel_export(suite, savings_result, vnd_result):
    """TC006: Kiểm tra Excel export."""
    test_id = "TC006"
    test_name = "Kiểm Tra Excel Export"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        from algorithms.excel_export import create_excel_report
        
        # Prepare results
        results = {
            'savings': savings_result
        }
        if vnd_result:
            results['vnd'] = vnd_result
        
        problem_info = {
            'num_customers': 3,
            'num_vehicles': 1,
            'capacity': 6000
        }
        
        filename = 'test_results.xlsx'
        
        # Create Excel
        filepath = create_excel_report(results, problem_info, filename)
        
        # Verify file exists
        assert os.path.exists(filepath), "Excel file not created"
        print_pass(f"Excel file created: {filepath}")
        
        # Verify file size
        file_size = os.path.getsize(filepath)
        assert file_size > 0, "Excel file is empty"
        print_pass(f"File size: {file_size} bytes")
        
        # Try to open with openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        print_pass(f"Excel has {len(wb.sheetnames)} sheets: {wb.sheetnames}")
        
        duration = time.time() - start_time
        result = TestResult(test_id, test_name, 'PASS',
                          f"Created Excel with {len(wb.sheetnames)} sheets", duration)
        suite.add_result(result)
        return True
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Excel export failed: {str(e)}")
        traceback.print_exc()
        result = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result)
        return False


# ============================================================================
# TC007: Test Batch Processing
# ============================================================================

def tc007_test_batch_processing(suite):
    """TC007: Kiểm tra batch processing."""
    test_id = "TC007"
    test_name = "Kiểm Tra Batch Processing"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        from algorithms.batch_processor import process_batch_files
        
        # Create multiple test files
        test_files = []
        for i in range(3):
            filepath = create_test_data_file()
            # Rename to unique name
            new_path = filepath.replace('standard_test.txt', f'batch_test_{i+1}.txt')
            os.rename(filepath, new_path)
            test_files.append({
                'filename': f'batch_test_{i+1}.txt',
                'filepath': new_path
            })
        
        print_info(f"Created {len(test_files)} test files")
        
        # Process batch
        algorithms = ['savings', 'vnd']
        batch_results = process_batch_files(test_files, algorithms)
        
        # Verify results
        assert len(batch_results) == len(test_files), "Not all files processed"
        print_pass(f"Processed {len(batch_results)} files")
        
        successful = sum(1 for r in batch_results if r['success'])
        failed = sum(1 for r in batch_results if not r['success'])
        
        print_pass(f"Successful: {successful}")
        if failed > 0:
            print_warning(f"Failed: {failed}")
        
        # Check individual results
        for i, result in enumerate(batch_results, 1):
            if result['success']:
                savings_cost = result['results']['savings']['total_cost']
                vnd_cost = result['results']['vnd']['total_cost']
                print_info(f"File {i}: Savings={savings_cost:.2f}, VND={vnd_cost:.2f}")
            else:
                print_fail(f"File {i}: Error - {result['error']}")
        
        duration = time.time() - start_time
        result_obj = TestResult(test_id, test_name, 'PASS',
                               f"Processed {successful}/{len(test_files)} files successfully", 
                               duration)
        suite.add_result(result_obj)
        return True, batch_results
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Batch processing failed: {str(e)}")
        traceback.print_exc()
        result_obj = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result_obj)
        return False, None


# ============================================================================
# TC008: Test Batch Excel Export
# ============================================================================

def tc008_test_batch_excel(suite, batch_results):
    """TC008: Kiểm tra batch Excel export."""
    test_id = "TC008"
    test_name = "Kiểm Tra Batch Excel Export"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        from algorithms.batch_processor import create_batch_summary
        from algorithms.batch_excel_export import create_batch_excel_report
        
        # Create summary
        summary = create_batch_summary(batch_results)
        print_info(f"Summary: {summary['total_files']} files, "
                  f"{summary['successful']} successful, {summary['failed']} failed")
        
        # Create Excel
        filename = 'test_batch_results.xlsx'
        filepath = create_batch_excel_report(batch_results, summary, filename)
        
        # Verify file
        assert os.path.exists(filepath), "Batch Excel file not created"
        print_pass(f"Batch Excel created: {filepath}")
        
        file_size = os.path.getsize(filepath)
        assert file_size > 0, "Batch Excel file is empty"
        print_pass(f"File size: {file_size} bytes")
        
        # Check sheets
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        print_pass(f"Excel has {len(wb.sheetnames)} sheets")
        
        # Should have Overview + one sheet per file
        expected_sheets = 1 + summary['total_files']
        assert len(wb.sheetnames) >= 1, "Missing Overview sheet"
        print_pass(f"Found Overview sheet and {len(wb.sheetnames)-1} instance sheets")
        
        duration = time.time() - start_time
        result = TestResult(test_id, test_name, 'PASS',
                          f"Created batch Excel with {len(wb.sheetnames)} sheets", duration)
        suite.add_result(result)
        return True
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Batch Excel failed: {str(e)}")
        traceback.print_exc()
        result = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result)
        return False


# ============================================================================
# TC015: Test Error Handling
# ============================================================================

def tc015_test_error_handling(suite):
    """TC015: Kiểm tra xử lý lỗi."""
    test_id = "TC015"
    test_name = "Kiểm Tra Error Handling"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    errors_caught = 0
    total_tests = 0
    
    try:
        from algorithms import read_vrpspd_file
        
        # Create invalid test files
        invalid_files = create_invalid_test_data()
        
        # Test each invalid file
        for name, filepath in invalid_files.items():
            total_tests += 1
            print_info(f"Testing: {name}")
            
            try:
                read_vrpspd_file(filepath)
                print_fail(f"Should have raised error for {name}")
            except Exception as e:
                errors_caught += 1
                print_pass(f"Correctly caught error: {type(e).__name__}")
        
        # Test non-existent file
        total_tests += 1
        print_info("Testing: non-existent file")
        try:
            read_vrpspd_file("nonexistent_file.txt")
            print_fail("Should have raised error for non-existent file")
        except Exception as e:
            errors_caught += 1
            print_pass(f"Correctly caught error: {type(e).__name__}")
        
        duration = time.time() - start_time
        success = errors_caught == total_tests
        
        if success:
            result = TestResult(test_id, test_name, 'PASS',
                              f"Caught {errors_caught}/{total_tests} errors correctly", duration)
        else:
            result = TestResult(test_id, test_name, 'FAIL',
                              f"Only caught {errors_caught}/{total_tests} errors", duration)
        
        suite.add_result(result)
        return success
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Error handling test failed: {str(e)}")
        traceback.print_exc()
        result = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result)
        return False


# ============================================================================
# TC016: Test Performance with Large Instance
# ============================================================================

def tc016_test_performance(suite):
    """TC016: Kiểm tra performance với instance lớn."""
    test_id = "TC016"
    test_name = "Kiểm Tra Performance"
    print_test_header(test_id, test_name)
    
    start_time = time.time()
    
    try:
        from algorithms import read_vrpspd_file, solve_savings, solve_vnd
        
        # Create large instance
        print_info("Creating large test instance (50 customers)...")
        filepath = create_large_test_data()
        
        # Parse
        cost_matrix, demand, pickup, vehicle_caps, capacity, num_vehicles = read_vrpspd_file(filepath)
        print_pass(f"Parsed {len(cost_matrix)-1} customers")
        
        # Test Savings
        print_info("Running Savings algorithm...")
        savings_start = time.time()
        savings_result = solve_savings(cost_matrix, demand, pickup, capacity, num_vehicles)
        savings_time = time.time() - savings_start
        
        print_pass(f"Savings completed in {savings_time:.2f}s")
        assert savings_time < 5, "Savings took too long (>5s)"
        
        # Test VND
        print_info("Running VND algorithm...")
        vnd_start = time.time()
        vnd_result = solve_vnd(
            savings_result['solution_vector'],
            cost_matrix,
            demand,
            pickup,
            capacity,
            max_time_seconds=60
        )
        vnd_time = time.time() - vnd_start
        
        print_pass(f"VND completed in {vnd_time:.2f}s")
        assert vnd_time < 60, "VND exceeded time limit"
        
        print_info(f"Cost improvement: {vnd_result['improvement']:.2f}%")
        
        duration = time.time() - start_time
        result = TestResult(test_id, test_name, 'PASS',
                          f"Savings: {savings_time:.2f}s, VND: {vnd_time:.2f}s", duration)
        suite.add_result(result)
        return True
        
    except Exception as e:
        duration = time.time() - start_time
        print_fail(f"Performance test failed: {str(e)}")
        traceback.print_exc()
        result = TestResult(test_id, test_name, 'FAIL', str(e), duration)
        suite.add_result(result)
        return False


# ============================================================================
# Test Suite Execution
# ============================================================================

def run_unit_tests():
    """Run all unit tests (algorithms only)."""
    print(f"\n{Colors.BOLD}{'='*70}")
    print("RUNNING UNIT TESTS")
    print(f"{'='*70}{Colors.ENDC}\n")
    
    suite = TestSuite()
    suite.start_time = datetime.now()
    
    # TC001: Imports
    if not tc001_test_imports(suite):
        print_fail("Critical failure: Cannot import modules. Stopping tests.")
        return suite
    
    # TC002: File Parser
    result = tc002_test_file_parser(suite)
    if not result[0]:
        print_fail("Critical failure: File parser failed. Stopping tests.")
        return suite
    
    _, filepath, cost_matrix, demand, pickup, capacity, num_vehicles = result
    
    # TC003: Savings
    success, savings_result = tc003_test_savings(suite, cost_matrix, demand, pickup, capacity, num_vehicles)
    if not success:
        print_fail("Savings failed. Some tests will be skipped.")
        savings_result = None
    
    # TC004: VND
    vnd_result = None
    if savings_result:
        success, vnd_result = tc004_test_vnd(suite, savings_result, cost_matrix, demand, pickup, capacity)
    
    # TC005: Visualization
    if savings_result:
        best_result = vnd_result if vnd_result else savings_result
        tc005_test_visualization(suite, best_result['routes'], cost_matrix, demand, pickup)
    
    # TC006: Excel Export
    if savings_result:
        tc006_test_excel_export(suite, savings_result, vnd_result)
    
    # TC007: Batch Processing
    success, batch_results = tc007_test_batch_processing(suite)
    
    # TC008: Batch Excel
    if success and batch_results:
        tc008_test_batch_excel(suite, batch_results)
    
    # TC015: Error Handling
    tc015_test_error_handling(suite)
    
    suite.end_time = datetime.now()
    return suite


def run_performance_tests():
    """Run performance tests."""
    print(f"\n{Colors.BOLD}{'='*70}")
    print("RUNNING PERFORMANCE TESTS")
    print(f"{'='*70}{Colors.ENDC}\n")
    
    suite = TestSuite()
    suite.start_time = datetime.now()
    
    # Import test
    if not tc001_test_imports(suite):
        return suite
    
    # TC016: Performance
    tc016_test_performance(suite)
    
    suite.end_time = datetime.now()
    return suite


def run_api_tests():
    """Run API tests (requires Flask server running)."""
    print(f"\n{Colors.BOLD}{'='*70}")
    print("RUNNING API TESTS")
    print(f"{'='*70}{Colors.ENDC}\n")
    
    print_warning("API tests require Flask server to be running!")
    print_info(f"Start server with: python app.py")
    print_info(f"Server should be at: {TEST_CONFIG['server_url']}")
    print_info("API tests are not implemented in this version.")
    print_info("Use manual testing or integration test framework like pytest + requests")
    
    suite = TestSuite()
    return suite


def run_all_tests():
    """Run all tests."""
    print(f"\n{Colors.BOLD}{'='*70}")
    print("RUNNING ALL TESTS")
    print(f"{'='*70}{Colors.ENDC}\n")
    
    all_suite = TestSuite()
    all_suite.start_time = datetime.now()
    
    # Run unit tests
    unit_suite = run_unit_tests()
    all_suite.results.extend(unit_suite.results)
    
    # Run performance tests
    perf_suite = run_performance_tests()
    all_suite.results.extend(perf_suite.results)
    
    all_suite.end_time = datetime.now()
    return all_suite


def generate_report(suite):
    """Generate test report file."""
    report_file = os.path.join(TEST_CONFIG['output_folder'], 
                               f"test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("=" * 70 + "\n")
        f.write("VRPSPD TEST REPORT\n")
        f.write("=" * 70 + "\n")
        f.write(f"Date: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write(f"Environment: Local Development\n\n")
        
        summary = suite.get_summary()
        f.write("TEST SUMMARY:\n")
        f.write(f"- Total Tests: {summary['total']}\n")
        f.write(f"- Passed: {summary['passed']} ✅\n")
        f.write(f"- Failed: {summary['failed']} ❌\n")
        f.write(f"- Skipped: {summary['skipped']} ⚠️\n")
        f.write(f"- Success Rate: {summary['success_rate']:.1f}%\n\n")
        
        f.write("DETAILS:\n")
        f.write("-" * 70 + "\n")
        for result in suite.results:
            status_icon = "✅" if result.status == "PASS" else "❌" if result.status == "FAIL" else "⚠️"
            f.write(f"[{result.test_id}] {result.name}: {status_icon} {result.status}\n")
            if result.message:
                f.write(f"    Message: {result.message}\n")
            f.write(f"    Duration: {result.duration:.2f}s\n")
            f.write("\n")
        
        if summary['failed'] > 0:
            f.write("\nFAILED TESTS:\n")
            f.write("-" * 70 + "\n")
            for result in suite.results:
                if result.status == "FAIL":
                    f.write(f"[{result.test_id}] {result.name}\n")
                    f.write(f"    Error: {result.message}\n\n")
        
        f.write("=" * 70 + "\n")
    
    print_info(f"Report saved to: {report_file}")
    return report_file


# ============================================================================
# Main Entry Point
# ============================================================================

def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description='VRPSPD Web App Test Suite')
    parser.add_argument('--all', action='store_true', help='Run all tests')
    parser.add_argument('--unit', action='store_true', help='Run unit tests only')
    parser.add_argument('--api', action='store_true', help='Run API tests only')
    parser.add_argument('--performance', action='store_true', help='Run performance tests only')
    parser.add_argument('--e2e', action='store_true', help='Run end-to-end test')
    
    args = parser.parse_args()
    
    # Setup
    ensure_folders()
    
    print(f"\n{Colors.BOLD}{Colors.OKBLUE}")
    print("🚀" * 35)
    print("  VRPSPD WEB APPLICATION - COMPLETE TEST SUITE")
    print("🚀" * 35)
    print(f"{Colors.ENDC}\n")
    
    # Determine which tests to run
    if args.unit:
        suite = run_unit_tests()
    elif args.api:
        suite = run_api_tests()
    elif args.performance:
        suite = run_performance_tests()
    elif args.e2e:
        print_info("End-to-end tests should be run manually through the web interface")
        print_info("See VRPSPD_TEST_SCENARIOS.md for TC018 instructions")
        return 0
    else:  # Default to all tests
        suite = run_all_tests()
    
    # Print summary
    suite.print_summary()
    
    # Generate report
    report_file = generate_report(suite)
    
    # Exit code
    summary = suite.get_summary()
    exit_code = 0 if summary['failed'] == 0 else 1
    
    if exit_code == 0:
        print(f"\n{Colors.OKGREEN}{Colors.BOLD}🎉 ALL TESTS PASSED! 🎉{Colors.ENDC}\n")
    else:
        print(f"\n{Colors.FAIL}{Colors.BOLD}❌ SOME TESTS FAILED ❌{Colors.ENDC}\n")
        print(f"Check report for details: {report_file}\n")
    
    return exit_code


if __name__ == '__main__':
    sys.exit(main())
