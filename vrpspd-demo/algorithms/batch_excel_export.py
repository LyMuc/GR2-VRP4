"""
Batch Excel Export Module for VRPSPD
Xuất kết quả batch processing ra file Excel master
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os


def create_batch_excel_report(batch_results, summary, filename=None):
    """
    Tạo file Excel master cho batch processing results.
    
    Args:
        batch_results (list): List of results cho từng file
        summary (dict): Summary statistics
        filename (str): Tên file (optional)
        
    Returns:
        str: Đường dẫn file Excel đã tạo
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"VRPSPD_Batch_Results_{timestamp}.xlsx"
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Summary
    create_batch_summary_sheet(wb, batch_results, summary)
    
    # Sheet 2: Comparison Table
    create_batch_comparison_sheet(wb, batch_results)
    
    # Sheet 3: Best Results
    create_batch_best_results_sheet(wb, batch_results)
    
    # Save file
    filepath = os.path.join('static', 'uploads', filename)
    wb.save(filepath)
    
    return filepath


def create_batch_summary_sheet(wb, batch_results, summary):
    """Tạo sheet tổng hợp batch results."""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = 'VRPSPD BATCH PROCESSING REPORT'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30
    
    # Timestamp
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Summary Statistics
    row = 4
    ws[f'A{row}'] = 'SUMMARY STATISTICS'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    summary_data = [
        ['Total Files:', summary.get('total_files', 0)],
        ['Successful:', summary.get('successful', 0)],
        ['Failed:', summary.get('failed', 0)],
        ['Average Improvement:', f"{summary.get('avg_improvement', 0):.2f}%"],
    ]
    
    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if 'Improvement' in label:
            ws[f'B{row}'].font = Font(bold=True, color='006100')
            ws[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        row += 1
    
    # Cost Summary
    row += 1
    ws[f'A{row}'] = 'COST SUMMARY'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = 'Total Savings Cost:'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = summary.get('total_cost_savings', 0)
    
    row += 1
    ws[f'A{row}'] = 'Total VND Cost:'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = summary.get('total_cost_vnd', 0)
    
    row += 1
    ws[f'A{row}'] = 'Total Savings:'
    ws[f'A{row}'].font = Font(bold=True)
    savings_amount = summary.get('total_cost_savings', 0) - summary.get('total_cost_vnd', 0)
    ws[f'B{row}'] = savings_amount
    ws[f'B{row}'].font = Font(bold=True, color='006100')
    ws[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # Apply borders
    apply_borders(ws, 1, 1, row, 5)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18


def create_batch_comparison_sheet(wb, batch_results):
    """Tạo sheet so sánh chi tiết từng file."""
    ws = wb.create_sheet("Comparison")
    
    # Title
    ws['A1'] = 'DETAILED COMPARISON'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 25
    
    # Headers
    row = 3
    headers = ['#', 'File', 'Customers', 'Vehicles', 'Savings Cost', 'Savings Time (s)', 'VND Cost', 'VND Time (s)', 'Improvement (%)', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row += 1
    for idx, result in enumerate(batch_results, 1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = result.get('filename', '-')
        
        if result.get('success'):
            problem_info = result.get('problem_info', {})
            savings = result.get('results', {}).get('savings', {})
            vnd = result.get('results', {}).get('vnd', {})
            
            ws[f'C{row}'] = problem_info.get('num_customers', '-')
            ws[f'D{row}'] = problem_info.get('num_vehicles', '-')
            ws[f'E{row}'] = savings.get('total_cost', 0)
            ws[f'F{row}'] = savings.get('computation_time', 0)
            ws[f'F{row}'].number_format = '0.000'
            ws[f'G{row}'] = vnd.get('total_cost', 0)
            ws[f'H{row}'] = vnd.get('computation_time', 0)
            ws[f'H{row}'].number_format = '0.000'
            ws[f'I{row}'] = vnd.get('improvement', 0)
            ws[f'I{row}'].number_format = '0.00'
            ws[f'J{row}'] = 'Success'
            ws[f'J{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
            # Highlight improvement
            if vnd.get('improvement', 0) > 0:
                ws[f'I{row}'].font = Font(bold=True, color='006100')
                ws[f'I{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        else:
            ws[f'C{row}'] = '-'
            ws[f'D{row}'] = '-'
            ws[f'E{row}'] = '-'
            ws[f'F{row}'] = '-'
            ws[f'G{row}'] = '-'
            ws[f'H{row}'] = '-'
            ws[f'I{row}'] = '-'
            ws[f'J{row}'] = 'Failed'
            ws[f'J{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws[f'J{row}'].font = Font(color='9C0006')
        
        # Alternate row colors
        if row % 2 == 0:
            for col in range(1, 11):
                if ws.cell(row, col).fill.start_color.rgb != 'C6EFCE' and ws.cell(row, col).fill.start_color.rgb != 'FFC7CE':
                    ws.cell(row, col).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        row += 1
    
    # Apply borders
    apply_borders(ws, 3, 1, row - 1, 10)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 12


def create_batch_best_results_sheet(wb, batch_results):
    """Tạo sheet hiển thị top best results."""
    ws = wb.create_sheet("Best Results")
    
    # Title
    ws['A1'] = 'TOP IMPROVEMENTS'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 25
    
    # Sort by improvement
    successful_results = [r for r in batch_results if r.get('success') and 'vnd' in r.get('results', {})]
    successful_results.sort(key=lambda x: x['results']['vnd'].get('improvement', 0), reverse=True)
    
    # Headers
    row = 3
    headers = ['Rank', 'File', 'Savings Cost', 'VND Cost', 'Cost Saved', 'Improvement (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data (top 10)
    row += 1
    for idx, result in enumerate(successful_results[:10], 1):
        savings = result['results']['savings']
        vnd = result['results']['vnd']
        cost_saved = savings['total_cost'] - vnd['total_cost']
        
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = result['filename']
        ws[f'C{row}'] = savings['total_cost']
        ws[f'D{row}'] = vnd['total_cost']
        ws[f'E{row}'] = cost_saved
        ws[f'F{row}'] = vnd.get('improvement', 0)
        ws[f'F{row}'].number_format = '0.00'
        
        # Gold, silver, bronze medals
        if idx == 1:
            medal_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        elif idx == 2:
            medal_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        elif idx == 3:
            medal_fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
        else:
            medal_fill = None
        
        if medal_fill:
            ws[f'A{row}'].fill = medal_fill
            ws[f'A{row}'].font = Font(bold=True)
        
        # Highlight improvement
        ws[f'F{row}'].font = Font(bold=True, color='006100')
        ws[f'F{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        row += 1
    
    # Apply borders
    apply_borders(ws, 3, 1, row - 1, 6)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18


def apply_borders(ws, start_row, start_col, end_row, end_col):
    """Áp dụng borders cho một vùng cells."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = thin_border
            if ws.cell(row, col).value is not None:
                ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
